import pandas
from openpyxl import load_workbook
import re 

file_path = 'your file path.xlsx'
wb = load_workbook(file_path ,data_only = True)#data_only保证读取为数值，防止读取到的为公式
ws = wb['your_work_sheet']#设定读取的工作表，一般为工作表一

column_index= 1#此处设定列数索引，此处设定为第一列
column_range = ws.iter_cols(min_col = column_index+1, max_col = column_index+1)#此处设定希望读取的列数，可通过改变index值来改变，这里读取的就是第二列到第二列，即读取第二列


data =  []
for row in column_range:
    column_data = [cell.value for cell in row]
    data.extend(column_data)
cell = ws.cell(row, column_index+1)#此处设定为单元格的行数与列数，此处为第row行，第2列

#获取单元格的数字格式：
number_format = cell.number_format

#使用正则表达式来读取单元格中小数的位数：
decimal_places = 0 #现将小数位数赋为0 
match = re.search(r'\.(\d+)',number_format)#设立正则读取小数点后位数
if match:
    decimal_places = len(match.group(1))#读取第一个捕获组，即（\d+）中的字符串长度
rounded_data = [round(value, decimal_places)if  isinstance(value,(int,float)) else None for value in data]#. 利用round函数对data中的数据进行保留设定的小数位数

#将其保存为DataFrame的形式
column_name = ws.cell(row, column = column_index+1).value#读取第row行第2列的名称
df = pandas.DataFrame({column_name:rounded_data})
print(df)